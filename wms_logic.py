"""
wms_logic.py — WMS-BATA
Lectura con openpyxl read_only (sin parsear estilos) para evitar SIGKILL en Render.
"""
import pandas as pd
import openpyxl


def leer_hoja_readonly(filepath, sheet_name):
    """Lee una hoja de Excel con openpyxl en modo read_only (liviano en memoria)."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Hojas disponibles: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    header = rows[0]
    data   = rows[1:]
    return pd.DataFrame(data, columns=header)


def fix_cols(df):
    df.columns = [
        c.encode('latin1').decode('utf-8', errors='replace') if isinstance(c, str) else c
        for c in df.columns
    ]
    return df


def sku_multiplier(sku):
    sku = str(sku).strip()
    if len(sku) == 15:
        try:
            return int(sku[10:12])
        except Exception:
            return 1
    return 1


def map_piso_vectorizado(df):
    area  = df['AREA'].astype(str).str.strip().str.upper()
    ubic  = df['UBICACION'].astype(str).str.strip().str.upper()
    nivel = pd.Series('1ER PISO', index=df.index)
    nivel = nivel.mask(ubic.str.startswith('CDBUFFER-C'), 'BUFFER-C')
    nivel = nivel.mask(area.isin(['PISO','PARED','DIS','ACT','SIN']), 'OMITIR')
    nivel = nivel.mask(area == 'MZN02', '2DO PISO')
    nivel = nivel.mask(area == 'MZN03', '3ER PISO')
    nivel = nivel.mask(area == 'MZN04', '4TO PISO')
    return nivel


def area_asignada_vectorizado(df):
    etq      = df['ETIQUETA'].astype(str).str.strip().str.upper()
    nivel    = df['NIVEL_BEST']
    resultado = nivel.where(nivel.notna() & (nivel != ''), 'SIN STOCK')
    resultado = resultado.mask(etq.isin(['INSUMOS','UNIFORMES']), etq)
    return resultado


def puede_atender_vectorizado(df):
    disp      = df['DISP_CONV_TOT']
    pend      = df['PEND_CONVERTIDO']
    resultado = pd.Series('PARCIAL', index=df.index)
    resultado = resultado.mask(disp >= pend, 'SÍ')
    resultado = resultado.mask(pend <= 0, 'NO')
    resultado = resultado.mask(disp.isna() | (disp == 0), 'NO')
    return resultado


def procesar_wms(filepath):
    """
    Lee el Excel (hojas PEND, STOCK, RUTA Y PRIORIDAD) con openpyxl read_only.
    Retorna (df_full, stk_valid).
    """
    # Verificar hojas disponibles
    wb_check = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    hojas = wb_check.sheetnames
    wb_check.close()

    requeridas = ['PEND', 'STOCK', 'RUTA Y PRIORIDAD']
    faltantes  = [h for h in requeridas if h not in hojas]
    if faltantes:
        raise ValueError(f"Faltan hojas: {', '.join(faltantes)}. Encontradas: {', '.join(hojas)}")

    df_pend_raw = fix_cols(leer_hoja_readonly(filepath, 'PEND'))
    df_stock_raw= fix_cols(leer_hoja_readonly(filepath, 'STOCK'))
    df_ruta_raw = fix_cols(leer_hoja_readonly(filepath, 'RUTA Y PRIORIDAD'))

    # ── PEND ──
    pend = pd.DataFrame()
    pend['N_ORDEN']       = df_pend_raw.iloc[:, 0]
    pend['ESTADO_ORDEN']  = df_pend_raw.iloc[:, 1]
    pend['SKU']           = df_pend_raw.iloc[:, 2].astype(str).str.strip()
    pend['CANT_SOL']      = pd.to_numeric(df_pend_raw.iloc[:, 3], errors='coerce').fillna(0)
    pend['CANT_ASIG']     = pd.to_numeric(df_pend_raw.iloc[:, 4], errors='coerce').fillna(0)
    pend['TDA_DESTINO']   = df_pend_raw.iloc[:, 5]
    pend['DESCRIPCION']   = df_pend_raw.iloc[:, 6]

    # ── STOCK ──
    stk = pd.DataFrame()
    stk['AREA']          = df_stock_raw.iloc[:, 0].astype(str).str.strip()
    stk['SKU']           = df_stock_raw.iloc[:, 1].astype(str).str.strip()
    stk['DESC_STK']      = df_stock_raw.iloc[:, 2]
    stk['UBICACION']     = df_stock_raw.iloc[:, 3].astype(str).str.strip()
    stk['CANT_ACTUAL']   = pd.to_numeric(df_stock_raw.iloc[:, 4], errors='coerce').fillna(0)
    stk['CANT_ASIG_STK'] = pd.to_numeric(df_stock_raw.iloc[:, 5], errors='coerce').fillna(0)

    # ── RUTA ──
    cols_req = ['LIMA / PROV','ruta','tda','Orden','semana','DIA CORREO','NOMBR','ETIQUETA']
    faltan   = [c for c in cols_req if c not in df_ruta_raw.columns]
    if faltan:
        raise ValueError(f"Faltan columnas en RUTA Y PRIORIDAD: {', '.join(faltan)}")

    ruta = pd.DataFrame()
    ruta['LIMA_PROV']     = df_ruta_raw['LIMA / PROV']
    ruta['RUTA']          = df_ruta_raw['ruta']
    ruta['TDA']           = df_ruta_raw['tda']
    ruta['ORDEN']         = df_ruta_raw['Orden']
    ruta['SEMANA']        = df_ruta_raw['semana'].astype(str)
    ruta['DIA_CORREO']    = df_ruta_raw['DIA CORREO'].astype(str)
    ruta['NOMBRE_TIENDA'] = df_ruta_raw['NOMBR']
    ruta['ETIQUETA']      = df_ruta_raw['ETIQUETA']

    # ── CÁLCULOS PENDIENTE ──
    pend['PENDIENTE']       = pend['CANT_SOL'] - pend['CANT_ASIG']
    pend['MULT']            = pend['SKU'].apply(sku_multiplier)
    pend['PEND_CONVERTIDO'] = pend['PENDIENTE'] * pend['MULT']

    # ── CÁLCULOS STOCK ──
    stk['DISPONIBLE']  = stk['CANT_ACTUAL'] - stk['CANT_ASIG_STK']
    stk['MULT']        = stk['SKU'].apply(sku_multiplier)
    stk['DISP_CONV']   = stk['DISPONIBLE'] * stk['MULT']
    stk['NIVEL']       = map_piso_vectorizado(stk)

    stk_valid = stk[stk['NIVEL'] != 'OMITIR'].copy()

    disp_conv_tot = (stk_valid.groupby('SKU')['DISP_CONV'].sum()
                     .reset_index().rename(columns={'DISP_CONV':'DISP_CONV_TOT'}))

    nivel_disp = (stk_valid.groupby(['SKU','NIVEL'])['DISP_CONV']
                  .sum().reset_index().rename(columns={'DISP_CONV':'DISP_NIVEL'}))
    if len(nivel_disp):
        best_nivel = (nivel_disp.loc[nivel_disp.groupby('SKU')['DISP_NIVEL'].idxmax(),
                                     ['SKU','NIVEL']].rename(columns={'NIVEL':'NIVEL_BEST'}))
    else:
        best_nivel = pd.DataFrame(columns=['SKU','NIVEL_BEST'])

    disp_tot = disp_conv_tot.merge(best_nivel, on='SKU', how='left')

    # ── JOIN PEND → RUTA ──
    ruta_u = ruta.drop_duplicates('ORDEN')[
        ['ORDEN','LIMA_PROV','RUTA','NOMBRE_TIENDA','SEMANA','DIA_CORREO','ETIQUETA']
    ].copy()
    pend['N_ORDEN'] = pd.to_numeric(pend['N_ORDEN'], errors='coerce')
    ruta_u['ORDEN'] = pd.to_numeric(ruta_u['ORDEN'], errors='coerce')

    df_full = pend.merge(ruta_u, left_on='N_ORDEN', right_on='ORDEN', how='left')
    df_full['SEMANA_DIA'] = df_full['SEMANA'].fillna('') + ' ' + df_full['DIA_CORREO'].fillna('')

    # ── JOIN → STOCK ──
    df_full = df_full.merge(disp_tot, on='SKU', how='left')

    # ── ÁREA Y PUEDE ATENDER (vectorizado) ──
    df_full['AREA_ASIGNADA'] = area_asignada_vectorizado(df_full)
    df_full['PUEDE_ATENDER'] = puede_atender_vectorizado(df_full)

    return df_full, stk_valid


def resumen_kpis(df_full):
    return {
        "total_lineas":    int(len(df_full)),
        "pend_conv_total": int(df_full['PEND_CONVERTIDO'].sum()),
        "si":      int((df_full['PUEDE_ATENDER'] == 'SÍ').sum()),
        "parcial": int((df_full['PUEDE_ATENDER'] == 'PARCIAL').sum()),
        "no":      int((df_full['PUEDE_ATENDER'] == 'NO').sum()),
    }
