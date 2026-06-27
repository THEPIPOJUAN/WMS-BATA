from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from functools import wraps
import json, os, io

app = Flask(__name__)
app.secret_key = "wms_bata_2026_secretkey"

USERS = {
    "jhuamani": {"password": "Bata2026", "role": "admin"}
}

DATA_FILE = "data.json"

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"prescrito": {}, "personal": [], "semana_actual": "SEM 23"}

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify({"error": "Sin permisos de edicion"}), 403
        return f(*args, **kwargs)
    return decorated

@app.route("/")
@login_required
def index():
    return render_template("index.html", user=session.get("user"), role=session.get("role"))

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        password = request.form.get("password", "").strip()
        if usuario in USERS and USERS[usuario]["password"] == password:
            session["user"] = usuario
            session["role"] = USERS[usuario]["role"]
            return redirect(url_for("index"))
        error = "Usuario o contrasena incorrectos"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/api/prescrito", methods=["GET"])
@login_required
def get_prescrito():
    data = load_data()
    return jsonify(data.get("prescrito", {}))

@app.route("/api/prescrito", methods=["POST"])
@login_required
@admin_required
def save_prescrito():
    data = load_data()
    data["prescrito"] = request.json
    save_data(data)
    return jsonify({"ok": True})

@app.route("/api/prescrito/excel")
@login_required
def download_prescrito_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = load_data()
    presData = data.get("prescrito", {})
    DIAS = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES"]
    COLS = ["accesorios", "calzado", "insumos", "muestras", "materiales", "doble"]
    LABELS = ["ACCESORIOS", "CALZADO", "INSUMOS", "MUESTRAS", "MATERIALES", "DOBLE TRAMO"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prescrito"

    rojo_oscuro = "991B1B"
    rojo_medio  = "7F1D1D"
    rojo_claro  = "FEE2E2"
    blanco      = "FFFFFF"
    gris_fila   = "FEF9F9"
    negro       = "1F2937"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def cs(cell, bg, fg="FFFFFF", bold=False, size=11, align="center"):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        cell.border = border

    totals = {c: sum(int(presData.get(d, {}).get(c, 0) or 0) for d in DIAS) for c in COLS}
    grand = sum(totals.values())

    kpi_titles = ["TOTAL GENERAL", "CALZADO", "INSUMOS", "ACCESORIOS", "DOBLE TRAMO"]
    kpi_vals   = [grand, totals["calzado"], totals["insumos"], totals["accesorios"], totals["doble"]]
    kpi_colors = ["991B1B", "991B1B", "854D0E", "1F2937", "166534"]
    kpi_starts = [1, 3, 5, 7, 9]

    for i, (title, val, color, col_start) in enumerate(zip(kpi_titles, kpi_vals, kpi_colors, kpi_starts)):
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+1)
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+1)
        c1 = ws.cell(row=1, column=col_start, value=title)
        cs(c1, "000000", "FFFFFF", bold=False, size=9)
        c2 = ws.cell(row=2, column=col_start, value=val)
        cs(c2, "000000", color, bold=True, size=14)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8

    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = "PRESCRITO"
    cs(c, rojo_oscuro, blanco, bold=True, size=13)
    ws.row_dimensions[4].height = 24

    headers = ["ATENCION"] + LABELS + ["TOTAL GENERAL"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        cs(c, rojo_medio, blanco, bold=True, size=10)
    ws.row_dimensions[5].height = 20

    for row_idx, dia in enumerate(DIAS, 6):
        vals = [int(presData.get(dia, {}).get(col, 0) or 0) for col in COLS]
        total = sum(vals)
        bg = gris_fila if row_idx % 2 == 0 else blanco
        c = ws.cell(row=row_idx, column=1, value=dia)
        cs(c, bg, negro, bold=True, size=11, align="left")
        for ci, v in enumerate(vals, 2):
            cc = ws.cell(row=row_idx, column=ci, value=v)
            cs(cc, bg, negro, size=11)
        ct = ws.cell(row=row_idx, column=8, value=total)
        cs(ct, rojo_claro, rojo_oscuro, bold=True, size=11)
        ws.row_dimensions[row_idx].height = 20

    total_row = 6 + len(DIAS)
    col_totals = [sum(int(presData.get(d, {}).get(c, 0) or 0) for d in DIAS) for c in COLS]
    grand_total = sum(col_totals)
    ct = ws.cell(row=total_row, column=1, value="TOTAL")
    cs(ct, rojo_oscuro, blanco, bold=True, size=11, align="left")
    for ci, v in enumerate(col_totals, 2):
        cc = ws.cell(row=total_row, column=ci, value=v)
        cs(cc, rojo_oscuro, blanco, bold=True, size=11)
    cg = ws.cell(row=total_row, column=8, value=grand_total)
    cs(cg, rojo_oscuro, blanco, bold=True, size=12)
    ws.row_dimensions[total_row].height = 22

    anchos = [16, 13, 13, 13, 13, 13, 14, 16]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="prescrito_wms_bata.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/personal", methods=["GET"])
@login_required
def get_personal():
    data = load_data()
    return jsonify(data.get("personal", []))


@app.route("/api/personal", methods=["POST"])
@login_required
@admin_required
def save_personal():
    data = load_data()
    data["personal"] = request.json
    save_data(data)
    return jsonify({"ok": True})


@app.route("/api/personal/excel")
@login_required
def download_personal_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = load_data()
    personal = data.get("personal", [])
    semana = request.args.get("semana", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personal"

    rojo = "991B1B"
    rojo_m = "7F1D1D"
    blanco = "FFFFFF"
    negro = "111827"
    gris = "F9FAFB"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def cs(cell, bg, fg="FFFFFF", bold=False, size=10, align="center"):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = f"CONTROL DE PERSONAL — {semana}"
    cs(c, rojo, blanco, bold=True, size=13)
    ws.row_dimensions[1].height = 24

    headers = ["N°", "DNI", "APELLIDOS Y NOMBRE", "AREA", "PUESTO", "ENCARGADO",
               "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "AREA"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        cs(c, rojo_m, blanco, bold=True, size=9)
    ws.row_dimensions[2].height = 18

    COLORES_VAL = {
        "SI":         ("D1FAE5", "065F46"),
        "NO":         ("FEE2E2", "991B1B"),
        "VACACIONES": ("FEF3C7", "92400E"),
        "SUSPENDIDO": ("EDE9FE", "5B21B6"),
        "D.M":        ("DBEAFE", "1E40AF"),
        "C.M":        ("DBEAFE", "1E40AF"),
        "CUMPLEAÑOS": ("DBEAFE", "1E40AF"),
    }

    for row_i, p in enumerate(personal, 3):
        bg = gris if row_i % 2 == 0 else blanco
        vals_fijos = [row_i-2, p.get("dni",""), p.get("nombre",""),
                      p.get("area",""), p.get("puesto",""), p.get("encargado","")]
        aligns = ["center","center","left","left","left","left"]
        for ci, (v, al) in enumerate(zip(vals_fijos, aligns), 1):
            cc = ws.cell(row=row_i, column=ci, value=v)
            cs(cc, bg, negro, size=10, align=al)

        dias_keys = [f"{semana}-{i}" for i in range(6)]
        for ci, dk in enumerate(dias_keys, 7):
            val = (p.get("dias") or {}).get(dk, "")
            col_bg, col_fg = COLORES_VAL.get(val.upper() if val else "", (bg, negro))
            cc = ws.cell(row=row_i, column=ci, value=val)
            cs(cc, col_bg, col_fg, bold=bool(val), size=10)

        cc = ws.cell(row=row_i, column=13, value=p.get("area_abrev",""))
        cs(cc, bg, negro, size=10)
        ws.row_dimensions[row_i].height = 18

    anchos = [5, 12, 28, 20, 20, 22, 10, 10, 10, 10, 10, 10, 12]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="personal_wms_bata.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/personal/importar", methods=["POST"])
@login_required
@admin_required
def importar_personal():
    import openpyxl
    from datetime import datetime, timedelta

    if 'archivo' not in request.files:
        return jsonify({"error": "No se envio archivo"}), 400

    archivo = request.files['archivo']
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active

    personal = []
    filas = list(ws.iter_rows(values_only=True))

    header_row = None
    fecha_cols = []
    for i, fila in enumerate(filas):
        fila_str = [str(c).upper() if c else '' for c in fila]
        if 'DNI' in fila_str:
            header_row = i
            for j, val in enumerate(fila):
                if isinstance(val, (int, float)) and 40000 < val < 60000:
                    fecha_cols.append((j, val))
            break

    if header_row is None:
        return jsonify({"error": "No se encontro cabecera DNI"}), 400

    def excel_date(serial):
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(serial))
        except:
            return None

    OMITIR = ['RENUNCIO','LI','DESPACHO','NO RETAIL','NOCHE',
              'PASO A LI','PASO A NO RETAIL','WEB','']

    contador = 0
    for fila in filas[header_row+1:]:
        if not fila or not fila[1]:
            continue
        dni      = str(fila[1]).strip() if fila[1] else ''
        nombre   = str(fila[2]).strip() if fila[2] else ''
        area     = str(fila[3]).strip() if fila[3] else ''
        puesto   = str(fila[4]).strip() if fila[4] else ''
        encargado = str(fila[5]).strip() if fila[5] else ''
        area_abrev = str(fila[-1]).strip() if fila[-1] else ''

        if not dni or not nombre:
            continue

        dias = {}
        for col_idx, serial in fecha_cols:
            fecha = excel_date(serial)
            if not fecha:
                continue
            year = fecha.year
            week = fecha.isocalendar()[1]
            dow  = fecha.weekday()
            dia_key = f"{year}-W{str(week).zfill(2)}-{dow}"
            val = str(fila[col_idx]).strip() if col_idx < len(fila) and fila[col_idx] else ''
            if val and val.upper() not in OMITIR:
                dias[dia_key] = val.upper()

        personal.append({
            "id": contador + 1,
            "dni": dni, "nombre": nombre, "area": area,
            "puesto": puesto, "encargado": encargado,
            "area_abrev": area_abrev, "dias": dias
        })
        contador += 1

    data = load_data()
    data["personal"] = personal
    save_data(data)
    return jsonify({"ok": True, "total": contador})


@app.route("/api/wms/historial", methods=["GET"])
@login_required
def get_wms_historial():
    data = load_data()
    return jsonify({"historial": data.get("wms_historial", [])})


@app.route("/api/wms/procesar", methods=["POST"])
@login_required
@admin_required
def procesar_wms_endpoint():
    from wms_logic import procesar_wms, resumen_kpis
    import datetime, json
    import pandas as pd

    if 'archivo' not in request.files:
        return jsonify({"error": "No se envio archivo"}), 400

    archivo = request.files['archivo']
    if not archivo.filename.lower().endswith('.xlsx'):
        return jsonify({"error": "El archivo debe ser .xlsx"}), 400

    try:
        df_full, stk_valid = procesar_wms(archivo)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Error procesando el archivo: {e}"}), 500

    kpis = resumen_kpis(df_full)

    # Guardar tabla en archivo separado (rotacion 0/1/2)
    data = load_data()
    historial = data.get("wms_historial", [])
    nuevo_idx = 0  # siempre ocupa el slot 0, los otros se corren

    COLS = ['N_ORDEN','SKU','DESCRIPCION','TDA_DESTINO','NOMBRE_TIENDA','LIMA_PROV','RUTA','PENDIENTE',
            'PEND_CONVERTIDO','DISP_CONV_TOT','AREA_ASIGNADA','PUEDE_ATENDER']
    sub = df_full[COLS].copy()
    sub['N_ORDEN']        = sub['N_ORDEN'].fillna('').astype(str)
    sub['SKU']            = sub['SKU'].astype(str)
    sub['DESCRIPCION']    = sub['DESCRIPCION'].fillna('').astype(str).str[:40]
    sub['TDA_DESTINO']    = sub['TDA_DESTINO'].fillna('').astype(str)
    sub['NOMBRE_TIENDA']  = sub['NOMBRE_TIENDA'].fillna('').astype(str).str[:40]
    sub['LIMA_PROV']      = sub['LIMA_PROV'].fillna('').astype(str)
    sub['RUTA']           = sub['RUTA'].fillna('').astype(str)
    sub['PENDIENTE']      = pd.to_numeric(sub['PENDIENTE'], errors='coerce').fillna(0).astype(int)
    sub['PEND_CONVERTIDO']= pd.to_numeric(sub['PEND_CONVERTIDO'], errors='coerce').fillna(0).astype(int)
    sub['DISP_CONV_TOT']  = pd.to_numeric(sub['DISP_CONV_TOT'], errors='coerce').fillna(0).astype(int)
    sub['AREA_ASIGNADA']  = sub['AREA_ASIGNADA'].fillna('').astype(str)
    sub['PUEDE_ATENDER']  = sub['PUEDE_ATENDER'].astype(str)

    # Renombrar slots: el nuevo ocupa wms_tabla_0, los anteriores suben de indice
    for i in range(min(len(historial), 1), -1, -1):
        src = f"wms_tabla_{i}.json"
        dst = f"wms_tabla_{i+1}.json"
        if os.path.exists(src):
            os.replace(src, dst)
    # Limpiar slot 3+ si existe
    if os.path.exists("wms_tabla_3.json"):
        os.remove("wms_tabla_3.json")

    with open("wms_tabla_0.json", "w", encoding="utf-8") as f:
        f.write(sub.to_json(orient='records', force_ascii=False))

    # Guardar stock para hoja STOCK VS PENDIENTE
    stk_sub = stk_valid[['SKU','AREA','CANT_ACTUAL']].copy()
    stk_sub['SKU'] = stk_sub['SKU'].astype(str)
    stk_sub['AREA'] = stk_sub['AREA'].astype(str)
    stk_sub['CANT_ACTUAL'] = pd.to_numeric(stk_sub['CANT_ACTUAL'], errors='coerce').fillna(0).astype(int)
    # Rotar slots stk igual que tabla
    for i in range(min(len(historial), 1), -1, -1):
        src = f"wms_stk_{i}.json"; dst = f"wms_stk_{i+1}.json"
        if os.path.exists(src): os.replace(src, dst)
    if os.path.exists("wms_stk_3.json"): os.remove("wms_stk_3.json")
    with open("wms_stk_0.json", "w", encoding="utf-8") as f:
        f.write(stk_sub.to_json(orient='records', force_ascii=False))

    nuevo_resultado = {
        "fecha": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "archivo": archivo.filename,
        "kpis": kpis,
        "tabla_file": "wms_tabla_0.json",
    }

    historial.insert(0, nuevo_resultado)
    data["wms_historial"] = historial[:3]
    save_data(data)

    return jsonify({"ok": True, "kpis": kpis})


@app.route("/api/wms/tabla", methods=["GET"])
@login_required
def get_wms_tabla():
    import json
    idx   = int(request.args.get("idx", 0))
    page  = int(request.args.get("page", 0))
    filtro_q = request.args.get("q", "").strip().upper()      # SÍ / PARCIAL / NO / ""
    filtro_r = request.args.get("ruta", "").strip().upper()   # nombre de ruta o ""
    PAGE_SIZE = 500

    tabla_file = f"wms_tabla_{idx}.json"
    if not os.path.exists(tabla_file):
        return jsonify({"error": "Tabla no disponible. Vuelve a procesar el Excel."}), 404

    with open(tabla_file, encoding="utf-8") as f:
        rows = json.load(f)

    # Filtrar
    if filtro_q:
        rows = [r for r in rows if r.get("PUEDE_ATENDER","") == filtro_q]
    if filtro_r:
        rows = [r for r in rows if filtro_r in str(r.get("RUTA","")).upper()]

    total = len(rows)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    chunk = rows[page * PAGE_SIZE : (page + 1) * PAGE_SIZE]

    return jsonify({
        "rows": chunk,
        "page": page,
        "total_pages": total_pages,
        "total_rows": total,
    })


@app.route("/api/wms/excel")
@login_required
def download_wms_excel():
    import json, io
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    idx = int(request.args.get("idx", 0))
    tabla_file = f"wms_tabla_{idx}.json"
    if not os.path.exists(tabla_file):
        return jsonify({"error": "Tabla no disponible. Vuelve a procesar el Excel."}), 404

    data = load_data()
    historial = data.get("wms_historial", [])
    if idx >= len(historial):
        return jsonify({"error": "Corrida no encontrada"}), 404

    meta  = historial[idx]
    kpis  = meta["kpis"]
    fecha = meta["fecha"]
    archivo_origen = meta["archivo"]

    with open(tabla_file, encoding="utf-8") as f:
        rows = json.load(f)

    ROJO="991B1B"; ROJO_M="7F1D1D"; ROJO_C="FEE2E2"
    VERDE="065F46"; VERDE_C="D1FAE5"
    AMARI="92400E"; AMARI_C="FEF3C7"
    NEGRO="111827"; BLANCO="FFFFFF"; GRIS="F9FAFB"
    thin = Side(style="thin", color="E5E7EB")
    brd  = Border(top=thin, bottom=thin, left=thin, right=thin)

    def cs(cell, bg, fg="FFFFFF", bold=False, size=10, align="center"):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        cell.border = brd

    wb = Workbook()

    # ── HOJA 1: RESUMEN ──
    ws1 = wb.active; ws1.title = "RESUMEN"
    kpi_data = [("TOTAL LINEAS",kpis['total_lineas'],NEGRO),("SI ATIENDE",kpis['si'],"166534"),
                ("PARCIAL",kpis['parcial'],"854D0E"),("NO ATIENDE",kpis['no'],ROJO),
                ("PEND.CONV",kpis['pend_conv_total'],"1E40AF")]
    col = 1
    for title, val, color in kpi_data:
        ws1.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+1)
        ws1.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+1)
        cs(ws1.cell(row=1,column=col,value=title), "000000","AAAAAA",size=8)
        cs(ws1.cell(row=2,column=col,value=val),   "000000",color,bold=True,size=14)
        col += 2
    ws1.row_dimensions[1].height=16; ws1.row_dimensions[2].height=26; ws1.row_dimensions[3].height=6
    ws1.merge_cells("A4:F4")
    cs(ws1.cell(row=4,column=1,value=f"RESUMEN POR RUTA — {archivo_origen} — {fecha}"),ROJO,BLANCO,bold=True,size=11)
    ws1.row_dimensions[4].height=20
    for i,h in enumerate(["TIENDA","NOMBRE","TOTAL","SÍ ATIENDE","PARCIAL","NO ATIENDE","% ATENCIÓN"],1):
        cs(ws1.cell(row=5,column=i,value=h),ROJO_M,BLANCO,bold=True,size=9)
    ws1.row_dimensions[5].height=18

    resumen = defaultdict(lambda:{'SI':0,'PARCIAL':0,'NO':0,'TOTAL':0,'NOMBRE':''})
    for r in rows:
        tda  = str(r.get('TDA_DESTINO','')).strip() or 'SIN TIENDA'
        nom  = str(r.get('NOMBRE_TIENDA','')).strip()
        q    = str(r.get('PUEDE_ATENDER','')).strip()
        resumen[tda]['NOMBRE'] = nom
        resumen[tda]['TOTAL'] += 1
        if q=='SÍ': resumen[tda]['SI'] += 1
        elif q=='PARCIAL': resumen[tda]['PARCIAL'] += 1
        else: resumen[tda]['NO'] += 1
    rutas = sorted(resumen.keys())

    for ri, tda in enumerate(rutas, 6):
        v = resumen[tda]
        pct = round((v['SI']+v['PARCIAL'])/v['TOTAL']*100,1) if v['TOTAL'] else 0
        bg = GRIS if ri%2==0 else BLANCO
        for ci,(val,al,fg) in enumerate(zip(
            [tda, v['NOMBRE'], v['TOTAL'], v['SI'], v['PARCIAL'], v['NO'], f"{pct}%"],
            ["center","left","center","center","center","center","center"],
            [NEGRO,NEGRO,NEGRO,VERDE,AMARI,ROJO,NEGRO]),1):
            cs(ws1.cell(row=ri,column=ci,value=val),bg,fg,size=10,align=al)
        ws1.row_dimensions[ri].height=18

    tr = 6+len(rutas)
    total_lin = len(rows)
    for ci,val in enumerate(["TOTAL","",total_lin,kpis['si'],kpis['parcial'],kpis['no'],
        f"{round((kpis['si']+kpis['parcial'])/total_lin*100,1)}%"],1):
        cs(ws1.cell(row=tr,column=ci,value=val),ROJO,BLANCO,bold=True,size=10)
    ws1.row_dimensions[tr].height=20
    for i,w in enumerate([10,28,10,12,10,12,12],1):
        ws1.column_dimensions[get_column_letter(i)].width=w

    # ── HOJA 2: DETALLE (sin estilos por fila para evitar SIGKILL) ──
    ws2 = wb.create_sheet("DETALLE")
    hdr_font = Font(bold=True, color=BLANCO, size=9, name="Calibri")
    hdr_fill = PatternFill("solid", fgColor=ROJO_M)
    hdr_alig = Alignment(horizontal="center", vertical="center")
    HDRS_DET = ["N° ORDEN","SKU","DESCRIPCIÓN","TIENDA","NOMBRE TIENDA","LIMA/PROV","RUTA","PENDIENTE","PEND.CONV","DISP.CONV","ÁREA","¿ATIENDE?"]
    for ci,h in enumerate(HDRS_DET,1):
        c = ws2.cell(row=1,column=ci,value=h)
        c.font=hdr_font; c.fill=hdr_fill; c.alignment=hdr_alig
    ws2.row_dimensions[1].height=16

    COLS_DET=['N_ORDEN','SKU','DESCRIPCION','TDA_DESTINO','NOMBRE_TIENDA','LIMA_PROV','RUTA','PENDIENTE','PEND_CONVERTIDO','DISP_CONV_TOT','AREA_ASIGNADA','PUEDE_ATENDER']
    for ri,r in enumerate(rows,2):
        for ci,col_key in enumerate(COLS_DET,1):
            ws2.cell(row=ri,column=ci,value=r.get(col_key,''))
    for i,w in enumerate([12,16,30,8,28,10,20,10,10,10,14,10],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    # ── HOJA 3: STOCK VS PENDIENTE ──
    ws3 = wb.create_sheet("STOCK VS PENDIENTE")
    ws3.merge_cells("A1:N1")
    cs(ws3.cell(row=1,column=1,value="STOCK VS PENDIENTE — SKUs SIN STOCK PARA OLA WMS"),ROJO,BLANCO,bold=True,size=13)
    ws3.row_dimensions[1].height=24

    # Cargar stock
    stk_file = f"wms_stk_{idx}.json"
    stk_rows = []
    if os.path.exists(stk_file):
        with open(stk_file, encoding="utf-8") as f:
            stk_rows = json.load(f)

    # SKUs sin stock (AREA_ASIGNADA == 'SIN STOCK')
    from collections import defaultdict
    sin_stock_pend = defaultdict(int)
    for r in rows:
        if str(r.get('AREA_ASIGNADA','')).strip() == 'SIN STOCK':
            sku = str(r.get('SKU',''))
            sin_stock_pend[sku] += int(r.get('PEND_CONVERTIDO',0) or 0)

    # Stock por SKU+AREA
    stk_area = defaultdict(lambda: defaultdict(int))
    for s in stk_rows:
        stk_area[str(s['SKU'])][str(s['AREA'])] += int(s.get('CANT_ACTUAL',0) or 0)

    # Ordenar por pendiente desc
    sin_stock_sorted = sorted(sin_stock_pend.items(), key=lambda x: -x[1])

    # Determinar max áreas para headers dinámicos
    max_areas = max((len(stk_area[sku]) for sku,_ in sin_stock_sorted), default=1) if sin_stock_sorted else 1
    max_areas = max(max_areas, 1)

    hdrs = ['SKU','CANT. PEND. CONV.']
    for i in range(1, max_areas+1):
        hdrs += [f'ÁREA {i}', f'QTY ÁREA {i}']

    for ci,h in enumerate(hdrs,1):
        c = ws3.cell(row=2, column=ci, value=h)
        cs(c, ROJO_M, BLANCO, bold=True, size=9)
    ws3.row_dimensions[2].height=16

    NARANJA_C = "FFF3E0"; NARANJA = "E65100"
    ROSA_C = "FFCDD2"; ROSA = "9C0006"

    if sin_stock_sorted:
        for ri, (sku, pconv) in enumerate(sin_stock_sorted, 3):
            bg = GRIS if ri%2==0 else BLANCO
            cs(ws3.cell(row=ri,column=1,value=sku), ROSA_C, NEGRO, size=9, align="left")
            cs(ws3.cell(row=ri,column=2,value=pconv), ROSA_C, ROSA, bold=True, size=9)
            areas = sorted(stk_area[sku].items(), key=lambda x: -x[1])
            if not areas:
                cs(ws3.cell(row=ri,column=3,value='SIN UBICACION'), "EEEEEE","757575",size=9)
                cs(ws3.cell(row=ri,column=4,value=0), "EEEEEE",NEGRO,size=9)
            else:
                for ai,(area_name,area_qty) in enumerate(areas):
                    col_a = 3 + ai*2; col_q = 4 + ai*2
                    cs(ws3.cell(row=ri,column=col_a,value=area_name), NARANJA_C, NEGRO, bold=True, size=9)
                    cs(ws3.cell(row=ri,column=col_q,value=area_qty), NARANJA_C, NARANJA, size=9)
            ws3.row_dimensions[ri].height=15
    else:
        c = ws3.cell(row=3,column=1,value='No hay SKUs sin stock ✓')
        cs(c, "D1FAE5", VERDE, bold=True, size=10, align="left")

    # Anchos automáticos hoja 3
    ws3.column_dimensions['A'].width=18
    ws3.column_dimensions['B'].width=18
    for i in range(3, 3+max_areas*2, 2):
        ws3.column_dimensions[get_column_letter(i)].width=14
        ws3.column_dimensions[get_column_letter(i+1)].width=10

    out=io.BytesIO(); wb.save(out); out.seek(0)
    fname = f"cruce_wms_{fecha[:10]}.xlsx"
    return send_file(out, download_name=fname, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
